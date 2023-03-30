import openpyxl

wb  = openpyxl.Workbook()  # create a workbook

ws = wb.create_sheet("MySheet")

# add some data
ws["A1"] = "Hello"
ws["A2"] = "Python"

ws.cell(row=4, column=1, value=20)
ws.cell(row=4, column=2, value=40)

ws.append([5,6,7,8,3,4,5])
ws.append(["hi","how","are","you"])

wb.save("test.xlsx")
