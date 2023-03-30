from openpyxl import load_workbook


wb = load_workbook("test.xlsx")

print(wb.sheetnames)

# select a sheet

ws = wb["MySheet"]

for row in ws.values:
    print(row)

# print the data in cell

print(ws["A1"].value)
